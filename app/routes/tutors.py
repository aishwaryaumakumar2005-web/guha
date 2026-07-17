from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, current_app
from flask_login import login_required
from app.extensions import db
from app.models import Tutor, Course
from app.helpers import admin_required, is_ajax_request
from app.forms import TutorForm
import tempfile

tutors_bp = Blueprint('tutors', __name__)

@tutors_bp.route('/tutors', methods=['GET', 'POST'])
@login_required
@admin_required
def list():
    if request.method == 'POST':
        form = TutorForm(request.form)
        if not form.validate():
            if is_ajax_request():
                return jsonify({"success": False, "errors": form.error_messages}), 400
            for msg in form.error_messages:
                flash(msg, 'danger')
            return redirect(url_for('tutors.list'))
        name = form.data.get('name', '').strip()
        email = form.data.get('email', '').strip()
        phone = form.data.get('phone', '').strip()
        specialization = form.data.get('specialization', '').strip()
        status = form.data.get('status', 'Active')
        selected_courses = [c for c in request.form.getlist('courses') if c]
        exists = Tutor.query.filter_by(email=email).first()
        if exists:
            message = f"Tutor with email '{email}' already exists!"
            if is_ajax_request():
                return jsonify({"success": False, "message": message}), 400
            flash(message, 'danger')
        else:
            new_tutor = Tutor(name=name, email=email, phone=phone, specialization=specialization, status=status)
            for c_id in selected_courses:
                course = Course.query.get(int(c_id))
                if course:
                    new_tutor.courses.append(course)
            db.session.add(new_tutor)
            db.session.commit()
            message = "Tutor added successfully!"
            if is_ajax_request():
                return jsonify({"success": True, "message": message}), 201
            flash(message, "success")
        return redirect(url_for('tutors.list'))
    all_tutors = Tutor.query.all()
    all_courses = Course.query.all()
    return render_template('tutors.html', tutors=all_tutors, courses=all_courses)

@tutors_bp.route('/tutors/edit/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit(id):
    tutor = Tutor.query.get_or_404(id)
    form = TutorForm(request.form)
    if not form.validate():
        if is_ajax_request():
            return jsonify({"success": False, "errors": form.error_messages}), 400
        for msg in form.error_messages:
            flash(msg, 'danger')
        return redirect(url_for('tutors.list'))
    new_email = form.data.get('email', '').strip()
    if new_email != tutor.email:
        exists = Tutor.query.filter_by(email=new_email).first()
        if exists:
            flash(f"Email '{new_email}' is already in use by another tutor.", 'danger')
            return redirect(url_for('tutors.list'))
    tutor.name = form.data.get('name', '').strip()
    tutor.email = new_email
    tutor.phone = form.data.get('phone', '').strip()
    tutor.specialization = form.data.get('specialization', '').strip()
    tutor.status = form.data.get('status', 'Active')
    tutor.courses = []
    for c_id in (c for c in request.form.getlist('courses') if c):
        course = Course.query.get(int(c_id))
        if course:
            tutor.courses.append(course)
    db.session.commit()
    message = "Tutor details updated!"
    if is_ajax_request():
        return jsonify({"success": True, "message": message}), 200
    flash(message, "success")
    return redirect(url_for('tutors.list'))

@tutors_bp.route('/tutors/delete/<int:id>')
@login_required
@admin_required
def delete(id):
    tutor = Tutor.query.get_or_404(id)
    db.session.delete(tutor)
    db.session.commit()
    message = "Tutor record removed!"
    if is_ajax_request():
        return jsonify({"success": True, "message": message}), 200
    flash(message, "success")
    return redirect(url_for('tutors.list'))

@tutors_bp.route('/api/tutors/import-excel', methods=['POST'])
@login_required
@admin_required
def import_excel():
    from werkzeug.utils import secure_filename
    from openpyxl import load_workbook
    import os
    if 'excel_file' not in request.files:
        return jsonify({"success": False, "errors": ["No file uploaded"]}), 400
    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({"success": False, "errors": ["No file selected"]}), 400
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"success": False, "errors": ["Invalid file format"]}), 400
    ai_validation = request.form.get('ai_validation', 'true').lower() == 'true'
    auto_course_mapping = request.form.get('auto_course_mapping', 'false').lower() == 'true'
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            file.save(tmp_file.name)
            tmp_file_path = tmp_file.name
        workbook = load_workbook(tmp_file_path)
        sheet = workbook.active
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[1]]
        column_map = {}
        for idx, header in enumerate(headers):
            if 'name' in header:
                column_map['name'] = idx
            elif 'email' in header:
                column_map['email'] = idx
            elif 'phone' in header or 'mobile' in header:
                column_map['phone'] = idx
            elif 'specialization' in header or 'special' in header:
                column_map['specialization'] = idx
            elif 'status' in header:
                column_map['status'] = idx
            elif 'course' in header:
                column_map['courses'] = idx
        required_columns = ['name', 'email', 'phone']
        missing_columns = [col for col in required_columns if col not in column_map]
        if missing_columns:
            return jsonify({"success": False, "errors": [f"Missing required columns: {', '.join(missing_columns)}"]}), 400
        tutors_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            tutor_data = {
                'name': str(row[column_map['name']]).strip() if column_map['name'] < len(row) else '',
                'email': str(row[column_map['email']]).strip() if column_map['email'] < len(row) else '',
                'phone': str(row[column_map['phone']]).strip() if column_map['phone'] < len(row) else '',
                'specialization': str(row[column_map['specialization']]).strip() if 'specialization' in column_map and column_map['specialization'] < len(row) else '',
                'status': str(row[column_map['status']]).strip() if 'status' in column_map and column_map['status'] < len(row) else 'Active',
                'courses': str(row[column_map['courses']]).strip() if 'courses' in column_map and column_map['courses'] < len(row) else ''
            }
            tutors_data.append(tutor_data)
        os.unlink(tmp_file_path)
        if not tutors_data:
            return jsonify({"success": False, "errors": ["No data found in Excel file"]}), 400
        ai_engine = current_app.ai_engine
        if ai_validation:
            validation_result = ai_engine.validate_tutor_data(tutors_data)
        else:
            validation_result = {"valid": True, "errors": [], "warnings": [], "suggestions": [], "enriched_data": tutors_data}
        if not validation_result["valid"]:
            return jsonify({"success": False, "errors": validation_result.get("errors", [])}), 400
        imported_count = 0
        skipped_count = 0
        all_courses = Course.query.all()
        all_courses_lower = {c.code.lower(): c for c in all_courses}
        existing_emails = set(email for (email,) in db.session.query(Tutor.email).all())
        for tutor_data in validation_result["enriched_data"]:
            if tutor_data['email'] in existing_emails:
                skipped_count += 1
                continue
            new_tutor = Tutor(
                name=tutor_data['name'], email=tutor_data['email'], phone=tutor_data['phone'],
                specialization=tutor_data.get('specialization', ''), status=tutor_data.get('status', 'Active')
            )
            if tutor_data.get('courses'):
                course_codes = [c.strip() for c in tutor_data['courses'].split(',')]
                for code in course_codes:
                    course = all_courses_lower.get(code.lower())
                    if course:
                        new_tutor.courses.append(course)
            elif auto_course_mapping:
                suggested_courses = ai_engine.suggest_course_mapping(tutor_data, all_courses)
                for course_id in suggested_courses:
                    course = next((c for c in all_courses if c.id == course_id), None)
                    if course:
                        new_tutor.courses.append(course)
            db.session.add(new_tutor)
            existing_emails.add(tutor_data['email'])
            imported_count += 1
        db.session.commit()
        return jsonify({"success": True, "message": f"Successfully imported {imported_count} tutors. Skipped {skipped_count} duplicates.", "imported": imported_count, "skipped": skipped_count}), 200
    except Exception as e:
        return jsonify({"success": False, "errors": [f"Error processing file: {str(e)}"]}), 500
