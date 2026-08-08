from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, current_app, send_file
from flask_login import login_required, current_user
from app.extensions import db
from app.models import Student, Course, Tutor, student_courses
from app.helpers import admin_required, is_ajax_request, save_photo_data
from app.forms import StudentForm
import tempfile
import io

students_bp = Blueprint('students', __name__)

@students_bp.route('/students', methods=['GET', 'POST'])
@login_required
def list():
    # Handle POST (add student) - only for admin
    if request.method == 'POST':
        if current_user.role != 'Admin':
            if is_ajax_request():
                return jsonify({"success": False, "errors": ["Only admins can add students"]}), 403
            flash("Only admins can add students", 'danger')
            return redirect(url_for('students.list'))
        
        form = StudentForm(request.form)
        if not form.validate():
            if is_ajax_request():
                return jsonify({"success": False, "errors": form.error_messages}), 400
            for msg in form.error_messages:
                flash(msg, 'danger')
            return redirect(url_for('students.list'))
        name = form.data.get('name', '').strip()
        email = form.data.get('email', '').strip()
        phone = form.data.get('phone', '').strip()
        status = form.data.get('status', 'Active')
        selected_courses = [c for c in request.form.getlist('courses') if c]
        exists = Student.query.filter_by(email=email).first()
        if exists:
            message = f"Student with email '{email}' already exists!"
            if is_ajax_request():
                return jsonify({"success": False, "message": message}), 400
            flash(message, 'danger')
        else:
            photo_data = None
            photo_mime = None
            if 'photo' in request.files and request.files['photo'].filename:
                try:
                    photo_data, photo_mime = save_photo_data(request.files['photo'])
                except ValueError as e:
                    if is_ajax_request():
                        return jsonify({"success": False, "errors": [str(e)]}), 400
                    flash(str(e), 'warning')
            new_student = Student(name=name, email=email, phone=phone, status=status, photo_data=photo_data, photo_mime=photo_mime)
            for c_id in selected_courses:
                course = Course.query.get(int(c_id))
                if course:
                    new_student.courses.append(course)
            db.session.add(new_student)
            db.session.commit()
            message = "Student enrolled successfully!"
            if is_ajax_request():
                return jsonify({"success": True, "message": message}), 201
            flash(message, "success")
        return redirect(url_for('students.list'))
    
    # GET request - filter students based on user role
    course_filter = request.args.get('course_id', type=int)
    if current_user.role == 'Staff':
        tutor = Tutor.query.filter_by(email=current_user.email).first()
        if tutor:
            course_ids = [c.id for c in tutor.courses]
            student_subquery = db.session.query(student_courses.c.student_id).filter(
                student_courses.c.course_id.in_(course_ids)
            ).distinct()
            all_students = Student.query.filter(Student.id.in_(student_subquery)).all()
        else:
            all_students = []
    else:
        if course_filter:
            student_subquery = db.session.query(student_courses.c.student_id).filter(
                student_courses.c.course_id == course_filter
            ).distinct()
            all_students = Student.query.filter(Student.id.in_(student_subquery)).order_by(Student.id).all()
        else:
            all_students = Student.query.order_by(Student.id).all()
    
    all_courses = Course.query.all()
    return render_template('students.html', students=all_students, courses=all_courses, is_staff=(current_user.role == 'Staff'), selected_course_id=course_filter)


@students_bp.route('/api/students/export-excel')
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    course_filter = request.args.get('course_id', type=int)
    if current_user.role == 'Staff':
        tutor = Tutor.query.filter_by(email=current_user.email).first()
        if tutor:
            course_ids = [c.id for c in tutor.courses]
            student_subquery = db.session.query(student_courses.c.student_id).filter(
                student_courses.c.course_id.in_(course_ids)
            ).distinct()
            students = Student.query.filter(Student.id.in_(student_subquery)).order_by(Student.id).all()
        else:
            students = []
    else:
        if course_filter:
            student_subquery = db.session.query(student_courses.c.student_id).filter(
                student_courses.c.course_id == course_filter
            ).distinct()
            students = Student.query.filter(Student.id.in_(student_subquery)).order_by(Student.id).all()
        else:
            students = Student.query.order_by(Student.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    headers = ['ID', 'Full Name', 'Email', 'Phone', 'Courses Enrolled', 'Enrollment Date', 'Status']
    header_fill = PatternFill(start_color="0D2740", end_color="0D2740", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style='thin', color='B0C4DE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    banded_fill = PatternFill(start_color="EAF2FB", end_color="EAF2FB", fill_type="solid")

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    for r, s in enumerate(students, start=2):
        course_names = ', '.join([c.name for c in s.courses]) if s.courses else ''
        status = s.status or 'Active'
        ws.cell(row=r, column=1, value=s.roll_no or s.id).number_format = '@'
        ws.cell(row=r, column=2, value=s.name)
        ws.cell(row=r, column=3, value=s.email or '')
        phone = ws.cell(row=r, column=4, value=(s.phone or ''))
        phone.number_format = '@'
        ws.cell(row=r, column=5, value=course_names)
        if s.enrollment_date:
            dc = ws.cell(row=r, column=6, value=s.enrollment_date)
            dc.number_format = 'DD MMM YYYY'
        else:
            ws.cell(row=r, column=6, value='')
        ws.cell(row=r, column=7, value=status)
        if r % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = banded_fill
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = border
            if col != 6:
                cell.alignment = Alignment(vertical="center")

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 18 if col != 5 else 32
        ws.column_dimensions[letter].bestFit = True

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:{0}{1}".format(get_column_letter(len(headers)), max(ws.max_row, 1))
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = 'students_export.xlsx'
    if course_filter:
        course = Course.query.get(course_filter)
        if course:
            filename = "students_{0}.xlsx".format(course.code or course.name).replace(' ', '_')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@students_bp.route('/students/edit/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit(id):
    student = Student.query.get_or_404(id)
    form = StudentForm(request.form)
    if not form.validate():
        if is_ajax_request():
            return jsonify({"success": False, "errors": form.error_messages}), 400
        for msg in form.error_messages:
            flash(msg, 'danger')
        return redirect(url_for('students.list'))
    new_email = form.data.get('email', '').strip()
    if new_email != student.email:
        exists = Student.query.filter_by(email=new_email).first()
        if exists:
            flash(f"Email '{new_email}' is already in use by another student.", 'danger')
            return redirect(url_for('students.list'))
    student.name = form.data.get('name', '').strip()
    student.email = new_email
    student.phone = form.data.get('phone', '').strip()
    student.status = form.data.get('status', 'Active')
    if request.form.get('remove_photo'):
        student.photo_data = None
        student.photo_mime = None
    elif 'photo' in request.files and request.files['photo'].filename:
        try:
            student.photo_data, student.photo_mime = save_photo_data(request.files['photo'])
        except ValueError as e:
            if is_ajax_request():
                return jsonify({"success": False, "errors": [str(e)]}), 400
            flash(str(e), 'danger')
    student.courses = []
    for c_id in (c for c in request.form.getlist('courses') if c):
        course = Course.query.get(int(c_id))
        if course:
            student.courses.append(course)
    db.session.commit()
    message = "Student details updated!"
    if is_ajax_request():
        return jsonify({"success": True, "message": message}), 200
    flash(message, "success")
    return redirect(url_for('students.list'))

@students_bp.route('/students/delete/<int:id>')
@login_required
@admin_required
def delete(id):
    student = Student.query.get_or_404(id)
    db.session.delete(student)
    db.session.commit()
    message = "Student record deleted!"
    if is_ajax_request():
        return jsonify({"success": True, "message": message}), 200
    flash(message, "success")
    return redirect(url_for('students.list'))

@students_bp.route('/api/students/check-duplicate')
@login_required
def check_duplicate():
    email = request.args.get('email', '').strip().lower()
    phone = request.args.get('phone', '').strip()
    exclude_id = request.args.get('exclude_id', type=int)
    matches = []
    q = Student.query
    if exclude_id:
        q = q.filter(Student.id != exclude_id)
    if email:
        other = q.filter(db.func.lower(Student.email) == email).first()
        if other:
            matches.append({'field': 'email', 'value': other.email, 'name': other.name})
    if phone:
        other = q.filter(Student.phone == phone).first()
        if other:
            matches.append({'field': 'phone', 'value': other.phone, 'name': other.name})
    return jsonify({'duplicates': matches})

@students_bp.route('/api/students/import-excel', methods=['POST'])
@login_required
@admin_required
def import_excel():
    from werkzeug.utils import secure_filename
    from openpyxl import load_workbook
    import os
    if 'excel_file' not in request.files:
        return jsonify({"success": False, "errors": ["No file uploaded"]}), 400
    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({"success": False, "errors": ["No file selected"]}), 400
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"success": False, "errors": ["Invalid file format. Please upload .xlsx or .xls file"]}), 400
    ai_validation = request.form.get('ai_validation', 'true').lower() == 'true'
    auto_course_mapping = request.form.get('auto_course_mapping', 'false').lower() == 'true'
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            file.save(tmp_file.name)
            tmp_file_path = tmp_file.name
        workbook = load_workbook(tmp_file_path)
        sheet = workbook.active
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        headers = [str(h).strip().lower() if h else '' for h in headers]
        column_map = {}
        for idx, header in enumerate(headers):
            if 'name' in header:
                column_map['name'] = idx
            elif 'email' in header:
                column_map['email'] = idx
            elif 'phone' in header or 'mobile' in header:
                column_map['phone'] = idx
            elif 'status' in header:
                column_map['status'] = idx
            elif 'course' in header:
                column_map['courses'] = idx
        required_columns = ['name', 'email', 'phone']
        missing_columns = [col for col in required_columns if col not in column_map]
        if missing_columns:
            return jsonify({
                "success": False,
                "errors": [f"Missing required columns: {', '.join(missing_columns)}"],
                "suggestions": ["Ensure your Excel file has columns: Name, Email, Phone"]
            }), 400
        students_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            student_data = {
                'name': str(row[column_map['name']]).strip() if column_map['name'] < len(row) else '',
                'email': str(row[column_map['email']]).strip() if column_map['email'] < len(row) else '',
                'phone': str(row[column_map['phone']]).strip() if column_map['phone'] < len(row) else '',
                'status': str(row[column_map['status']]).strip() if 'status' in column_map and column_map['status'] < len(row) else 'Active',
                'courses': str(row[column_map['courses']]).strip() if 'courses' in column_map and column_map['courses'] < len(row) else ''
            }
            students_data.append(student_data)
        os.unlink(tmp_file_path)
        if not students_data:
            return jsonify({"success": False, "errors": ["No data found in Excel file"]}), 400
        ai_engine = current_app.ai_engine
        if ai_validation:
            validation_result = ai_engine.validate_student_data(students_data)
        else:
            validation_result = {"valid": True, "errors": [], "warnings": [], "suggestions": [], "enriched_data": students_data}
        if not validation_result["valid"]:
            return jsonify({
                "success": False, "errors": validation_result.get("errors", []),
                "warnings": validation_result.get("warnings", []),
                "suggestions": validation_result.get("suggestions", [])
            }), 400
        imported_count = 0
        skipped_count = 0
        all_courses = Course.query.all()
        all_courses_lower = {c.code.lower(): c for c in all_courses}
        existing_emails = set(email for (email,) in db.session.query(Student.email).all())
        for student_data in validation_result["enriched_data"]:
            if student_data['email'] in existing_emails:
                skipped_count += 1
                continue
            new_student = Student(
                name=student_data['name'], email=student_data['email'],
                phone=student_data['phone'], status=student_data.get('status', 'Active')
            )
            if student_data.get('courses'):
                course_codes = [c.strip() for c in student_data['courses'].split(',')]
                for code in course_codes:
                    course = all_courses_lower.get(code.lower())
                    if course:
                        new_student.courses.append(course)
            elif auto_course_mapping:
                suggested_courses = ai_engine.suggest_course_mapping(student_data, all_courses)
                for course_id in suggested_courses:
                    course = next((c for c in all_courses if c.id == course_id), None)
                    if course:
                        new_student.courses.append(course)
            db.session.add(new_student)
            existing_emails.add(student_data['email'])
            imported_count += 1
        db.session.commit()
        return jsonify({
            "success": True, "message": f"Successfully imported {imported_count} students. Skipped {skipped_count} duplicates.",
            "imported": imported_count, "skipped": skipped_count,
            "warnings": validation_result.get("warnings", [])
        }), 200
    except Exception as e:
        return jsonify({"success": False, "errors": [f"Error processing file: {str(e)}"]}), 500


@students_bp.route('/students/photo/<int:student_id>')
@login_required
def student_photo(student_id):
    student = Student.query.get_or_404(student_id)
    if not student.photo_data:
        return '', 404
    return send_file(io.BytesIO(student.photo_data), mimetype=student.photo_mime or 'image/jpeg')
