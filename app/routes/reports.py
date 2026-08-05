from datetime import datetime, date, timedelta
from io import BytesIO
from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, send_file, current_app
from flask_login import login_required, current_user
from app.extensions import db
from app.models import FeeRecord, Expense, ExpenseCategory, Course, Student, student_courses, Attendance, Tutor, OwnerFunding, Company, Account
from app.helpers import admin_required
from app.services.account_service import compute_account_summary
from app.services.payment_methods import PAYMENT_METHODS, classify_method

reports_bp = Blueprint('reports', __name__)

def filter_by_company_methods(query, model_attr, company_id):
    """Filter Expense/OwnerFunding query by payment methods belonging to the given company.
    
    Account names are the canonical payment method names (Cash, UPI - Ejaj Sir, etc.).
    We check if classify_method(record.payment_method) matches any account name 
    belonging to the selected company.
    """
    if not company_id:
        return query
    # Get canonical account names for this company (these ARE the payment method names)
    accounts = Account.query.filter_by(company_id=company_id, is_active=True).all()
    if not accounts:
        return query.filter(False)
    canonical_names = {acc.name for acc in accounts}  # e.g. {'Cash', 'UPI - Ejaj Sir', ...}
    # Get all distinct payment methods actually used in the DB
    observed = set()
    for (m,) in db.session.query(FeeRecord.payment_method).distinct().all():
        if m: observed.add(m)
    for (m,) in db.session.query(Expense.payment_method).distinct().all():
        if m: observed.add(m)
    for (m,) in db.session.query(OwnerFunding.method).distinct().all():
        if m: observed.add(m)
    # Find which observed methods classify to one of this company's account names
    matched = [m for m in observed if classify_method(m) in canonical_names]
    if not matched:
        return query.filter(False)
    return query.filter(db.func.coalesce(model_attr, '').in_(matched))


@reports_bp.route('/reports')
@login_required
def reports():
    # Staff: show simplified reports for their courses
    if current_user.role == 'Staff':
        tutor = Tutor.query.filter_by(email=current_user.email).first()
        if not tutor:
            return render_template('reports.html', tab='staff', today=date.today(), is_staff=True,
                staff_data={'students': [], 'courses': [], 'attendance_rate': 0, 'total_collected': 0, 'recent_fees': []})
        
        course_ids = [c.id for c in tutor.courses]
        student_subquery = db.session.query(student_courses.c.student_id).filter(
            student_courses.c.course_id.in_(course_ids)
        ).distinct()
        students = Student.query.filter(Student.id.in_(student_subquery)).all()
        
        # Get attendance data for staff's students
        today = date.today()
        thirty_days_ago = today - timedelta(days=30)
        attendance_records = Attendance.query.filter(
            Attendance.person_type == 'student',
            Attendance.person_id.in_([s.id for s in students]),
            Attendance.date >= thirty_days_ago
        ).all()
        
        # Calculate attendance rates
        total_records = len(attendance_records)
        present_records = sum(1 for r in attendance_records if r.status == 'Present')
        attendance_rate = (present_records / total_records * 100) if total_records > 0 else 0
        
        # Get fee data for staff's students
        fee_records = FeeRecord.query.filter(
            FeeRecord.student_id.in_([s.id for s in students])
        ).order_by(FeeRecord.payment_date.desc()).limit(50).all()
        
        total_collected = sum(r.amount_paid for r in fee_records)
        
        return render_template('reports.html', 
            tab='staff', 
            today=today,
            is_staff=True,
            staff_data={
                'students': students,
                'courses': tutor.courses,
                'attendance_rate': round(attendance_rate, 1),
                'total_collected': total_collected,
                'recent_fees': fee_records[:20]
            })

    today = date.today()
    tab = request.args.get('tab', 'income')
    filter_mode = request.args.get('filter_mode', 'monthly')
    filter_month = request.args.get('month', type=int) or today.month
    filter_year = request.args.get('year', type=int) or today.year
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    quick = request.args.get('quick', '').strip().lower()
    selected_company_id = request.args.get('company_id', type=int)

    companies = Company.query.filter_by(is_active=True).all()

    def _quick_range(key):
        if key == 'today':
            return today, today
        if key == 'this_week':
            return today - timedelta(days=today.weekday()), today
        if key == 'this_month':
            return date(today.year, today.month, 1), today
        if key == 'last_month':
            end = date(today.year, today.month, 1) - timedelta(days=1)
            return date(end.year, end.month, 1), end
        if key == 'this_quarter':
            q = (today.month - 1) // 3 + 1
            qs = (q - 1) * 3 + 1
            return date(today.year, qs, 1), today
        if key == 'ytd':
            return date(today.year, 1, 1), today
        if key == 'all':
            return date(2000, 1, 1), date(2100, 12, 31)
        return None

    qrange = _quick_range(quick) if quick else None
    if qrange:
        start_date, end_date = qrange
        filter_mode = 'custom'
        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d')
    elif filter_mode == 'custom' and start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        if filter_mode == 'quarterly':
            q = (filter_month - 1) // 3 + 1
            q_start = (q - 1) * 3 + 1
            start_date = date(filter_year, q_start, 1)
            if q == 4:
                end_date = date(filter_year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(filter_year, q_start + 3, 1) - timedelta(days=1)
        else:
            start_date = date(filter_year, filter_month, 1)
            if filter_month == 12:
                end_date = date(filter_year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(filter_year, filter_month + 1, 1) - timedelta(days=1)

    months_names = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    
    # Query builder helper for FeeRecord with company_id filtering
    fee_q = FeeRecord.query
    if selected_company_id:
        fee_q = fee_q.filter(FeeRecord.company_id == selected_company_id)

    total_income_query = db.session.query(db.func.sum(FeeRecord.amount_paid))
    if selected_company_id:
        total_income_query = total_income_query.filter(FeeRecord.company_id == selected_company_id)
    total_income = total_income_query.scalar() or 0.0

    fee_monthly_query = db.session.query(
        db.extract('month', FeeRecord.payment_date).label('m'),
        db.func.sum(FeeRecord.amount_paid).label('total')
    ).filter(db.extract('year', FeeRecord.payment_date) == filter_year)
    if selected_company_id:
        fee_monthly_query = fee_monthly_query.filter(FeeRecord.company_id == selected_company_id)
    fee_monthly_rows = fee_monthly_query.group_by(db.extract('month', FeeRecord.payment_date)).all()
    
    fee_monthly_map = {}
    for r in fee_monthly_rows:
        fee_monthly_map[int(r.m)] = float(r.total)
    income_monthly = [fee_monthly_map.get(m, 0.0) for m in range(1, 13)]
    fees_monthly = [{"month": months_names[m-1], "total": fee_monthly_map.get(m, 0.0)} for m in range(1, 13)]

    course_fee_query = db.session.query(
        student_courses.c.course_id, db.func.sum(FeeRecord.amount_paid).label('total')
    ).select_from(FeeRecord).join(Student).join(student_courses).join(Course, Course.id == student_courses.c.course_id).filter(
        FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date
    )
    if selected_company_id:
        course_fee_query = course_fee_query.filter(
            FeeRecord.company_id == selected_company_id,
            Course.company_id == selected_company_id
        )
    else:
        course_fee_query = course_fee_query.filter(Course.company_id == FeeRecord.company_id)
    course_fee_rows = course_fee_query.group_by(student_courses.c.course_id).all()
    course_fee_map = {r.course_id: float(r.total) for r in course_fee_rows}
    
    course_wise_income = []
    for course in Course.query.all():
        total = course_fee_map.get(course.id, 0.0)
        if total > 0:
            course_wise_income.append({"name": course.name, "code": course.code, "total": total})

    daily_collections = fee_q.filter(
        FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date
    ).order_by(FeeRecord.payment_date.desc()).all()

    expense_categories = ExpenseCategory.query.all()
    exp_monthly_query = db.session.query(
        db.extract('month', Expense.expense_date).label('m'),
        db.func.sum(Expense.amount).label('total')
    ).filter(db.extract('year', Expense.expense_date) == filter_year)
    exp_monthly_query = filter_by_company_methods(exp_monthly_query, Expense.payment_method, selected_company_id)
    exp_monthly_rows = exp_monthly_query.group_by(db.extract('month', Expense.expense_date)).all()
    exp_monthly_map = {int(r.m): float(r.total) for r in exp_monthly_rows}
    monthly_expense = [{"month": months_names[m-1], "total": exp_monthly_map.get(m, 0.0)} for m in range(1, 13)]
    
    cat_exp_query = db.session.query(
        Expense.category_id, db.func.count(Expense.id).label('cnt'), db.func.sum(Expense.amount).label('total')
    ).filter(Expense.expense_date >= start_date, Expense.expense_date <= end_date)
    cat_exp_query = filter_by_company_methods(cat_exp_query, Expense.payment_method, selected_company_id)
    cat_exp_rows = cat_exp_query.group_by(Expense.category_id).all()
    cat_exp_map = {r.category_id: {'total': float(r.total), 'count': r.cnt} for r in cat_exp_rows}
    category_wise_expense = []
    expense_cat_map = {cat.id: cat for cat in expense_categories}
    for cat in expense_categories:
        d = cat_exp_map.get(cat.id, {'total': 0.0, 'count': 0})
        if d['total'] > 0:
            category_wise_expense.append({"name": cat.name, "total": d['total'], "count": d['count']})
    
    summary_query = db.session.query(
        Expense.category_id, db.func.count(Expense.id).label('cnt'), db.func.sum(Expense.amount).label('total')
    ).filter(
        db.extract('month', Expense.expense_date) == filter_month,
        db.extract('year', Expense.expense_date) == filter_year
    )
    summary_query = filter_by_company_methods(summary_query, Expense.payment_method, selected_company_id)
    summary_rows = summary_query.group_by(Expense.category_id).all()
    summary_map = {r.category_id: {'total': float(r.total), 'count': r.cnt} for r in summary_rows}
    expense_summary = []
    for cat in expense_categories:
        s = summary_map.get(cat.id, {'total': 0.0, 'count': 0})
        expense_summary.append({"name": cat.name, "total": s['total'], "count": s['count']})
    
    recent_expenses_query = Expense.query.filter(
        Expense.expense_date >= start_date, Expense.expense_date <= end_date
    )
    recent_expenses_query = filter_by_company_methods(recent_expenses_query, Expense.payment_method, selected_company_id)
    recent_expenses_q = recent_expenses_query.order_by(Expense.expense_date.desc()).limit(20).all()
    recent_expenses = []
    exp_chart_labels = []
    exp_chart_amounts = []
    for e in reversed(recent_expenses_q):
        recent_expenses.append(e)
        exp_chart_labels.append(e.expense_date.strftime('%d %b'))
        exp_chart_amounts.append(e.amount)

    tot_inc_query = db.session.query(db.func.sum(FeeRecord.amount_paid)).filter(
        FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date
    )
    tot_gst_query = db.session.query(db.func.sum(FeeRecord.gst_amount)).filter(
        FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date
    )
    tot_tax_query = db.session.query(db.func.sum(FeeRecord.taxable_amount)).filter(
        FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date
    )
    if selected_company_id:
        tot_inc_query = tot_inc_query.filter(FeeRecord.company_id == selected_company_id)
        tot_gst_query = tot_gst_query.filter(FeeRecord.company_id == selected_company_id)
        tot_tax_query = tot_tax_query.filter(FeeRecord.company_id == selected_company_id)

    total_income_filtered = tot_inc_query.scalar() or 0.0
    total_gst_filtered = tot_gst_query.scalar() or 0.0
    total_taxable_filtered = tot_tax_query.scalar() or 0.0

    total_expense_query = db.session.query(db.func.sum(Expense.amount)).filter(
        Expense.expense_date >= start_date, Expense.expense_date <= end_date
    )
    total_expense_query = filter_by_company_methods(total_expense_query, Expense.payment_method, selected_company_id)
    total_expense_filtered = total_expense_query.scalar() or 0.0

    funding_monthly_query = db.session.query(
        db.extract('month', OwnerFunding.funding_date).label('m'),
        db.func.sum(OwnerFunding.amount).label('total')
    ).filter(db.extract('year', OwnerFunding.funding_date) == filter_year)
    funding_monthly_query = filter_by_company_methods(funding_monthly_query, OwnerFunding.method, selected_company_id)
    funding_monthly_rows = funding_monthly_query.group_by(db.extract('month', OwnerFunding.funding_date)).all()
    funding_monthly_map = {int(r.m): float(r.total) for r in funding_monthly_rows}
    funding_monthly = [funding_monthly_map.get(m, 0.0) for m in range(1, 13)]

    total_funding_query = db.session.query(db.func.sum(OwnerFunding.amount)).filter(
        OwnerFunding.funding_date >= start_date, OwnerFunding.funding_date <= end_date
    )
    total_funding_query = filter_by_company_methods(total_funding_query, OwnerFunding.method, selected_company_id)
    total_funding_filtered = total_funding_query.scalar() or 0.0

    net_balance = float(total_income_filtered) + float(total_funding_filtered) - float(total_expense_filtered)
    
    pl_monthly = []
    for m in range(1, 13):
        inc = fee_monthly_map.get(m, 0.0)
        exp = exp_monthly_map.get(m, 0.0)
        fund = funding_monthly_map.get(m, 0.0)
        pl_monthly.append({"month": months_names[m-1], "income": inc, "expense": exp, "funding": fund, "net": inc + fund - exp})

    # Per-company income & P&L breakdown
    company_pl = []
    comp_inc_rows = db.session.query(
        FeeRecord.company_id,
        db.func.sum(FeeRecord.amount_paid).label('inc'),
        db.func.sum(FeeRecord.taxable_amount).label('tax'),
        db.func.sum(FeeRecord.gst_amount).label('gst')
    ).filter(FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date).group_by(FeeRecord.company_id).all()
    
    comp_inc_map = {r.company_id: {'inc': float(r.inc or 0), 'tax': float(r.tax or 0), 'gst': float(r.gst or 0)} for r in comp_inc_rows}

    for c in companies:
        cdata = comp_inc_map.get(c.id, {'inc': 0.0, 'tax': 0.0, 'gst': 0.0})
        company_pl.append({
            'company': c,
            'income': cdata['inc'],
            'taxable': cdata['tax'],
            'gst': cdata['gst'],
        })

    cumulative_income = []
    running = 0.0
    for v in income_monthly:
        running += v
        cumulative_income.append(round(running, 2))

    pm_query = db.session.query(FeeRecord.payment_method, db.func.sum(FeeRecord.amount_paid))
    if selected_company_id:
        pm_query = pm_query.filter(FeeRecord.company_id == selected_company_id)
    payment_methods = pm_query.group_by(FeeRecord.payment_method).all()
    payment_labels = [p[0] for p in payment_methods]
    payment_data = [float(p[1]) for p in payment_methods]

    thirty_days_ago = today - timedelta(days=30)
    dt_query = db.session.query(
        FeeRecord.payment_date, db.func.sum(FeeRecord.amount_paid)
    ).filter(FeeRecord.payment_date >= thirty_days_ago)
    if selected_company_id:
        dt_query = dt_query.filter(FeeRecord.company_id == selected_company_id)
    daily_totals = dt_query.group_by(FeeRecord.payment_date).order_by(FeeRecord.payment_date).all()
    daily_labels = [d[0].strftime('%d %b') for d in daily_totals]
    daily_amounts = [float(d[1]) for d in daily_totals]

    net_trend = [p['net'] for p in pl_monthly]

    period_fee_q = FeeRecord.query.filter(
        FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date
    )
    if selected_company_id:
        period_fee_q = period_fee_q.filter(FeeRecord.company_id == selected_company_id)
    period_fee_records = period_fee_q.order_by(FeeRecord.payment_date.desc()).all()

    def get_payment_method_data(account_name):
        records = [r for r in period_fee_records if classify_method(r.payment_method) == account_name]
        total = sum(r.amount_paid for r in records)
        return {'total': total, 'records': records}

    payment_methods_report = {}
    for m in PAYMENT_METHODS:
        payment_methods_report[m] = get_payment_method_data(m)
    payment_methods_report['Others'] = get_payment_method_data('Others')

    tot_col_query = db.session.query(db.func.sum(FeeRecord.amount_paid)).filter(
        FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date
    )
    if selected_company_id:
        tot_col_query = tot_col_query.filter(FeeRecord.company_id == selected_company_id)
    total_collected_period = tot_col_query.scalar() or 0.0

    return render_template('reports.html', tab=tab, today=today, filter_mode=filter_mode,
        filter_month=filter_month, filter_year=filter_year,
        start_date_str=start_date_str or start_date.strftime('%Y-%m-%d'),
        end_date_str=end_date_str or end_date.strftime('%Y-%m-%d'),
        active_quick=quick, companies=companies, selected_company_id=selected_company_id,
        total_income=float(total_income), income_monthly=income_monthly, cumulative_income=cumulative_income,
        fees_monthly=fees_monthly, course_wise_income=course_wise_income, daily_collections=daily_collections,
        payment_labels=payment_labels, payment_data=payment_data, daily_labels=daily_labels, daily_amounts=daily_amounts,
        monthly_expense=monthly_expense, category_wise_expense=category_wise_expense, expense_summary=expense_summary,
        recent_expenses=recent_expenses, exp_chart_labels=exp_chart_labels, exp_chart_amounts=exp_chart_amounts,
        total_income_filtered=float(total_income_filtered), total_expense_filtered=float(total_expense_filtered),
        total_gst_filtered=float(total_gst_filtered), total_taxable_filtered=float(total_taxable_filtered),
        total_funding_filtered=float(total_funding_filtered), funding_monthly=funding_monthly,
        net_balance=net_balance, pl_monthly=pl_monthly, company_pl=company_pl, net_trend=net_trend,
        payment_methods_report=payment_methods_report, total_collected_period=float(total_collected_period),
        account_balances=(compute_account_summary() if current_user.role == 'Admin' else []))

@reports_bp.route('/reports/pdf')
@login_required
@admin_required
def report_pdf():
    from fpdf import FPDF
    today = date.today()
    tab = request.args.get('tab', 'income')
    filter_mode = request.args.get('filter_mode', 'monthly')
    filter_month = request.args.get('month', type=int) or today.month
    filter_year = request.args.get('year', type=int) or today.year
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    selected_company_id = request.args.get('company_id', type=int)

    company_obj = Company.query.get(selected_company_id) if selected_company_id else None

    if filter_mode == 'custom' and start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        if filter_mode == 'quarterly':
            q = (filter_month - 1) // 3 + 1
            q_start = (q - 1) * 3 + 1
            start_date = date(filter_year, q_start, 1)
            end_date = date(filter_year + 1, 1, 1) - timedelta(days=1) if q == 4 else date(filter_year, q_start + 3, 1) - timedelta(days=1)
        else:
            start_date = date(filter_year, filter_month, 1)
            end_date = date(filter_year + 1, 1, 1) - timedelta(days=1) if filter_month == 12 else date(filter_year, filter_month + 1, 1) - timedelta(days=1)
    
    period_label = f"{start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}"
    company_label = f"  |  Company: {company_obj.name}" if company_obj else "  |  Company: All"

    class ReportPDF(FPDF):
        def header(self):
            self.set_font('Helvetica', 'B', 16)
            self.set_text_color(47, 72, 88)
            self.cell(0, 10, 'Guha Academy - Computer Institute', align='C', new_x="LMARGIN", new_y="NEXT")
            self.set_font('Helvetica', '', 9)
            self.set_text_color(100, 100, 100)
            titles = {'income': 'Income Report', 'fees': 'Student Fees Report', 'expense': 'Expense Report', 'overall': 'Overall Report'}
            self.cell(0, 6, f'{titles.get(tab, "Report")}  |  Period: {period_label}{company_label}', align='C', new_x="LMARGIN", new_y="NEXT")
            self.ln(4)
            self.set_draw_color(217, 93, 57)
            self.set_line_width(0.5)
            self.line(10, self.get_y(), 200, self.get_y())
            self.ln(6)

        def footer(self):
            self.set_y(-15)
            self.set_font('Helvetica', 'I', 7)
            self.set_text_color(150, 150, 150)
            self.cell(0, 10, f'Generated: {today.strftime("%d %b %Y %I:%M %p")}  |  Page {self.page_no()}/{{nb}}', align='C')

        def section_title(self, title):
            self.set_font('Helvetica', 'B', 12)
            self.set_text_color(47, 72, 88)
            self.cell(0, 8, title, new_x="LMARGIN", new_y="NEXT")
            self.ln(2)

        def kpi_box(self, label, value, color=(107, 142, 35)):
            self.set_font('Helvetica', 'B', 12)
            self.set_text_color(*color)
            self.cell(60, 7, value, align='C')
            self.set_font('Helvetica', '', 7)
            self.set_text_color(100, 100, 100)
            self.cell(0, 7, label, align='C', new_x="LMARGIN", new_y="NEXT")
            self.ln(1)

        def table_header(self, cols, widths):
            self.set_font('Helvetica', 'B', 8)
            self.set_fill_color(47, 72, 88)
            self.set_text_color(255, 255, 255)
            for i, col in enumerate(cols):
                self.cell(widths[i], 6, col, border=1, align='C', fill=True)
            self.ln()

        def table_row(self, cols, widths, aligns=None):
            self.set_font('Helvetica', '', 8)
            self.set_text_color(50, 50, 50)
            for i, col in enumerate(cols):
                a = aligns[i] if aligns else 'C'
                self.cell(widths[i], 5, str(col), border=1, align=a)
            self.ln()

    pdf = ReportPDF()
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    if tab == 'income':
        tot_q = db.session.query(db.func.sum(FeeRecord.amount_paid))
        gst_q = db.session.query(db.func.sum(FeeRecord.gst_amount)).filter(FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date)
        tax_q = db.session.query(db.func.sum(FeeRecord.taxable_amount)).filter(FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date)
        if selected_company_id:
            tot_q = tot_q.filter(FeeRecord.company_id == selected_company_id)
            gst_q = gst_q.filter(FeeRecord.company_id == selected_company_id)
            tax_q = tax_q.filter(FeeRecord.company_id == selected_company_id)
        
        total_income = tot_q.scalar() or 0.0
        gst_tot = gst_q.scalar() or 0.0
        tax_tot = tax_q.scalar() or 0.0

        inc_q = db.session.query(
            db.extract('month', FeeRecord.payment_date).label('m'), db.func.sum(FeeRecord.amount_paid).label('total')
        ).filter(db.extract('year', FeeRecord.payment_date) == filter_year)
        if selected_company_id:
            inc_q = inc_q.filter(FeeRecord.company_id == selected_company_id)
        income_rows = inc_q.group_by(db.extract('month', FeeRecord.payment_date)).all()
        income_map = {int(r.m): float(r.total) for r in income_rows}
        income_monthly = [income_map.get(m, 0.0) for m in range(1, 13)]
        
        pdf.section_title('Income Summary')
        pdf.kpi_box('Total Income (All Time)', f'Rs. {total_income:,.2f}', (107, 142, 35))
        pdf.kpi_box(f'Period Taxable Income', f'Rs. {tax_tot:,.2f}', (47, 72, 88))
        pdf.kpi_box(f'Period GST Collected', f'Rs. {gst_tot:,.2f}', (70, 130, 180))
        pdf.ln(4)
        pdf.section_title(f'Monthly Income - {filter_year}')
        pdf.table_header(['Month', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'], [20] + [14]*12)
        pdf.table_row(['Income (Rs.)'] + [f'{v:,.0f}' for v in income_monthly], [20] + [14]*12, ['L'] + ['R']*12)

    elif tab == 'fees':
        fees_q = db.session.query(
            db.extract('month', FeeRecord.payment_date).label('m'), db.func.sum(FeeRecord.amount_paid).label('total')
        ).filter(db.extract('year', FeeRecord.payment_date) == filter_year)
        if selected_company_id:
            fees_q = fees_q.filter(FeeRecord.company_id == selected_company_id)
        fees_rows = fees_q.group_by(db.extract('month', FeeRecord.payment_date)).all()
        fees_map = {int(r.m): float(r.total) for r in fees_rows}
        fees_monthly = [fees_map.get(m, 0.0) for m in range(1, 13)]
        months_names = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        pdf.section_title('Monthly Fees Collection')
        pdf.table_header(['Month'] + months_names, [20] + [14]*12)
        pdf.table_row(['Amount (Rs.)'] + [f'{v:,.0f}' for v in fees_monthly], [20] + [14]*12, ['L'] + ['R']*12)
        pdf.ln(3)

        course_fee_q = db.session.query(
            student_courses.c.course_id, db.func.sum(FeeRecord.amount_paid).label('total')
        ).select_from(FeeRecord).join(Student).join(student_courses).join(Course, Course.id == student_courses.c.course_id).filter(
            FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date
        )
        if selected_company_id:
            course_fee_q = course_fee_q.filter(
                FeeRecord.company_id == selected_company_id,
                Course.company_id == selected_company_id
            )
        else:
            course_fee_q = course_fee_q.filter(Course.company_id == FeeRecord.company_id)
        course_fee_rows = course_fee_q.group_by(student_courses.c.course_id).all()
        course_fee_map = {r.course_id: float(r.total) for r in course_fee_rows}
        course_wise = [(course.name, course_fee_map.get(course.id, 0.0)) for course in Course.query.all() if course_fee_map.get(course.id, 0.0) > 0]
        if course_wise:
            pdf.section_title('Course-wise Income')
            pdf.table_header(['Course', 'Amount (Rs.)'], [140, 50])
            for name, total in course_wise:
                pdf.table_row([name, f'{total:,.2f}'], [140, 50], ['L', 'R'])
        pdf.ln(3)

        daily_q = FeeRecord.query.filter(
            FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date
        )
        if selected_company_id:
            daily_q = daily_q.filter(FeeRecord.company_id == selected_company_id)
        daily = daily_q.order_by(FeeRecord.payment_date.desc()).all()
        if daily:
            pdf.section_title('Daily Collections (with Company & GST)')
            pdf.table_header(['Date', 'Student', 'Company', 'Total', 'GST', 'Method'], [24, 40, 48, 25, 20, 33])
            for r in daily[:30]:
                c_name = (r.company.name if r.company else 'Unassigned')[:24]
                pdf.table_row([r.payment_date.strftime('%d %b %Y'), r.student.name[:18], c_name, f'{r.amount_paid:,.0f}', f'{r.gst_amount:,.0f}', r.payment_method], [24, 40, 48, 25, 20, 33], ['L', 'L', 'L', 'R', 'R', 'C'])

    elif tab == 'expense':
        expense_cats = ExpenseCategory.query.all()
        months_names = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        exp_q = db.session.query(
            db.extract('month', Expense.expense_date).label('m'), db.func.sum(Expense.amount).label('total')
        ).filter(db.extract('year', Expense.expense_date) == filter_year)
        exp_q = filter_by_company_methods(exp_q, Expense.payment_method, selected_company_id)
        exp_rows = exp_q.group_by(db.extract('month', Expense.expense_date)).all()
        exp_map = {int(r.m): float(r.total) for r in exp_rows}
        monthly_exp = [exp_map.get(m, 0.0) for m in range(1, 13)]
        pdf.section_title('Monthly Expense')
        pdf.table_header(['Month'] + months_names, [20] + [14]*12)
        pdf.table_row(['Amount (Rs.)'] + [f'{v:,.0f}' for v in monthly_exp], [20] + [14]*12, ['L'] + ['R']*12)
        pdf.ln(3)
        cat_exp_q = db.session.query(
            Expense.category_id, db.func.sum(Expense.amount).label('total')
        ).filter(Expense.expense_date >= start_date, Expense.expense_date <= end_date)
        cat_exp_q = filter_by_company_methods(cat_exp_q, Expense.payment_method, selected_company_id)
        cat_exp_rows = cat_exp_q.group_by(Expense.category_id).all()
        cat_exp_map = {r.category_id: float(r.total) for r in cat_exp_rows}
        cat_wise = [(cat.name, cat_exp_map.get(cat.id, 0.0)) for cat in expense_cats if cat_exp_map.get(cat.id, 0.0) > 0]
        if cat_wise:
            pdf.section_title('Category-wise Expense')
            pdf.table_header(['Category', 'Amount (Rs.)'], [140, 50])
            for name, total in cat_wise:
                pdf.table_row([name, f'{total:,.2f}'], [140, 50], ['L', 'R'])
        pdf.ln(3)
        pdf.section_title(f'Expense Summary - {months_names[filter_month-1]} {filter_year}')
        pdf.table_header(['Category', 'Count', 'Total (Rs.)'], [100, 30, 60])
        summary_q = db.session.query(
            Expense.category_id, db.func.count(Expense.id).label('cnt'), db.func.sum(Expense.amount).label('total')
        ).filter(db.extract('month', Expense.expense_date) == filter_month, db.extract('year', Expense.expense_date) == filter_year)
        summary_q = filter_by_company_methods(summary_q, Expense.payment_method, selected_company_id)
        summary_rows = summary_q.group_by(Expense.category_id).all()
        summary_map = {r.category_id: {'total': float(r.total), 'count': r.cnt} for r in summary_rows}
        for cat in expense_cats:
            s = summary_map.get(cat.id, {'total': 0.0, 'count': 0})
            pdf.table_row([cat.name, str(s['count']), f'{s["total"]:,.2f}'], [100, 30, 60], ['L', 'C', 'R'])

    elif tab == 'overall':
        inc_q = db.session.query(db.func.sum(FeeRecord.amount_paid)).filter(FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date)
        gst_q = db.session.query(db.func.sum(FeeRecord.gst_amount)).filter(FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date)
        if selected_company_id:
            inc_q = inc_q.filter(FeeRecord.company_id == selected_company_id)
            gst_q = gst_q.filter(FeeRecord.company_id == selected_company_id)
        
        total_income = inc_q.scalar() or 0.0
        total_gst = gst_q.scalar() or 0.0
        
        total_exp_q = db.session.query(db.func.sum(Expense.amount)).filter(Expense.expense_date >= start_date, Expense.expense_date <= end_date)
        total_exp_q = filter_by_company_methods(total_exp_q, Expense.payment_method, selected_company_id)
        total_expense = total_exp_q.scalar() or 0.0
        
        total_fund_q = db.session.query(db.func.sum(OwnerFunding.amount)).filter(OwnerFunding.funding_date >= start_date, OwnerFunding.funding_date <= end_date)
        total_fund_q = filter_by_company_methods(total_fund_q, OwnerFunding.method, selected_company_id)
        total_funding = total_fund_q.scalar() or 0.0
        
        net = float(total_income) + float(total_funding) - float(total_expense)
        
        pdf.section_title('Profit & Loss Summary')
        pdf.kpi_box('Total Income', f'Rs. {float(total_income):,.2f}', (107, 142, 35))
        pdf.kpi_box('GST Collected', f'Rs. {float(total_gst):,.2f}', (70, 130, 180))
        pdf.kpi_box('Total Expense', f'Rs. {float(total_expense):,.2f}', (192, 57, 43))
        pdf.ln(3)
        pdf.kpi_box('Net Balance', f"{'+' if net >= 0 else ''}Rs. {net:,.2f}", (47, 72, 88) if net >= 0 else (192, 57, 43))
        pdf.ln(4)

        months_names = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        pdf.section_title(f'Monthly P&L - {filter_year}')
        pdf.table_header(['Month', 'Income', 'Funding', 'Expense', 'Net', 'Status'], [30, 40, 40, 40, 40, 30])
        inc_m_q = db.session.query(db.extract('month', FeeRecord.payment_date).label('m'), db.func.sum(FeeRecord.amount_paid).label('total')).filter(db.extract('year', FeeRecord.payment_date) == filter_year)
        if selected_company_id:
            inc_m_q = inc_m_q.filter(FeeRecord.company_id == selected_company_id)
        inc_rows = inc_m_q.group_by(db.extract('month', FeeRecord.payment_date)).all()
        inc_map = {int(r.m): float(r.total) for r in inc_rows}
        
        exp_m_q = db.session.query(db.extract('month', Expense.expense_date).label('m'), db.func.sum(Expense.amount).label('total')).filter(db.extract('year', Expense.expense_date) == filter_year)
        exp_m_q = filter_by_company_methods(exp_m_q, Expense.payment_method, selected_company_id)
        exp_rows_ov = exp_m_q.group_by(db.extract('month', Expense.expense_date)).all()
        exp_map_ov = {int(r.m): float(r.total) for r in exp_rows_ov}
        
        fund_m_q = db.session.query(db.extract('month', OwnerFunding.funding_date).label('m'), db.func.sum(OwnerFunding.amount).label('total')).filter(db.extract('year', OwnerFunding.funding_date) == filter_year)
        fund_m_q = filter_by_company_methods(fund_m_q, OwnerFunding.method, selected_company_id)
        fund_rows_ov = fund_m_q.group_by(db.extract('month', OwnerFunding.funding_date)).all()
        fund_map_ov = {int(r.m): float(r.total) for r in fund_rows_ov}
        for m in range(1, 13):
            inc = inc_map.get(m, 0.0)
            exp = exp_map_ov.get(m, 0.0)
            fund = fund_map_ov.get(m, 0.0)
            n = inc + fund - exp
            status = 'Profit' if n > 0 else ('Loss' if n < 0 else 'Breakeven')
            pdf.table_row([months_names[m-1], f'{inc:,.2f}', f'{fund:,.2f}', f'{exp:,.2f}', f"{'+' if n >= 0 else ''}{n:,.2f}", status], [30, 40, 40, 40, 40, 30], ['C', 'R', 'R', 'R', 'R', 'C'])

    elif tab == 'payment_methods':
        all_q = FeeRecord.query.filter(FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date)
        if selected_company_id:
            all_q = all_q.filter(FeeRecord.company_id == selected_company_id)
        all_records = all_q.order_by(FeeRecord.payment_date.desc()).all()
        total_period = sum(r.amount_paid for r in all_records)
        def filter_pm(records, account_name):
            matched = [r for r in records if classify_method(r.payment_method) == account_name]
            return sum(r.amount_paid for r in matched), matched
        pdf.section_title('Payment Methods Report')
        pdf.kpi_box('Total Collected', f'Rs. {total_period:,.2f}', (47, 72, 88))
        pdf.ln(4)
        for label in PAYMENT_METHODS + ['Others']:
            total, records = filter_pm(all_records, label)
            if not records and label != 'Others':
                continue
            pdf.section_title(f'{label} - Rs. {total:,.2f}')
            if records:
                pdf.table_header(['Date', 'Student', 'Amount', 'Remarks'], [30, 55, 35, 70])
                for r in records[:30]:
                    pdf.table_row([r.payment_date.strftime('%d %b %Y'), r.student.name[:20], f'{r.amount_paid:,.0f}', (r.remarks or '')[:30]], [30, 55, 35, 70], ['L', 'L', 'R', 'L'])
            pdf.ln(2)

    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    filenames = {'income': 'Income_Report', 'fees': 'Student_Fees_Report', 'expense': 'Expense_Report', 'overall': 'Overall_Report', 'payment_methods': 'Payment_Methods_Report'}
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=f'{filenames.get(tab, "Report")}_{filter_year}_{filter_month}.pdf')

@reports_bp.route('/reports/excel')
@login_required
@admin_required
def report_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    today = date.today()
    tab = request.args.get('tab', 'income')
    filter_mode = request.args.get('filter_mode', 'monthly')
    filter_month = request.args.get('month', type=int) or today.month
    filter_year = request.args.get('year', type=int) or today.year
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    selected_company_id = request.args.get('company_id', type=int)

    if filter_mode == 'custom' and start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        if filter_mode == 'quarterly':
            q = (filter_month - 1) // 3 + 1
            q_start = (q - 1) * 3 + 1
            start_date = date(filter_year, q_start, 1)
            end_date = date(filter_year + 1, 1, 1) - timedelta(days=1) if q == 4 else date(filter_year, q_start + 3, 1) - timedelta(days=1)
        else:
            start_date = date(filter_year, filter_month, 1)
            end_date = date(filter_year + 1, 1, 1) - timedelta(days=1) if filter_month == 12 else date(filter_year, filter_month + 1, 1) - timedelta(days=1)
    
    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F4858', end_color='2F4858', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def write_sheet(ws, title, headers, rows):
        ws.title = title
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border
        for r, row in enumerate(rows, 2):
            for col, val in enumerate(row, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.border = thin_border
                c.alignment = Alignment(horizontal='right' if isinstance(val, (int, float)) else 'left')
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    if tab == 'income':
        inc_q = db.session.query(db.extract('month', FeeRecord.payment_date).label('m'), db.func.sum(FeeRecord.amount_paid).label('total')).filter(db.extract('year', FeeRecord.payment_date) == filter_year)
        tot_q = db.session.query(db.func.sum(FeeRecord.amount_paid))
        gst_q = db.session.query(db.func.sum(FeeRecord.gst_amount)).filter(FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date)
        tax_q = db.session.query(db.func.sum(FeeRecord.taxable_amount)).filter(FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date)
        if selected_company_id:
            inc_q = inc_q.filter(FeeRecord.company_id == selected_company_id)
            tot_q = tot_q.filter(FeeRecord.company_id == selected_company_id)
            gst_q = gst_q.filter(FeeRecord.company_id == selected_company_id)
            tax_q = tax_q.filter(FeeRecord.company_id == selected_company_id)
        
        inc_rows = inc_q.group_by(db.extract('month', FeeRecord.payment_date)).all()
        inc_map = {int(r.m): float(r.total) for r in inc_rows}
        income_monthly = [inc_map.get(m, 0.0) for m in range(1, 13)]
        total_all_time = tot_q.scalar() or 0
        gst_period = gst_q.scalar() or 0
        tax_period = tax_q.scalar() or 0

        months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        ws = wb.active
        write_sheet(ws, 'Monthly Income', ['Month'] + months, [['Income (Rs.)'] + income_monthly])
        ws2 = wb.create_sheet('Summary')
        write_sheet(ws2, 'Summary', ['Metric', 'Value'], [
            ['Total Income (All Time)', total_all_time],
            [f'Period Taxable Income', tax_period],
            [f'Period GST Collected', gst_period],
            [f'Total ({filter_year})', sum(income_monthly)],
            [f'Monthly Avg ({filter_year})', sum(income_monthly) / 12]
        ])

    elif tab == 'fees':
        months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        fees_q = db.session.query(db.extract('month', FeeRecord.payment_date).label('m'), db.func.sum(FeeRecord.amount_paid).label('total')).filter(db.extract('year', FeeRecord.payment_date) == filter_year)
        if selected_company_id:
            fees_q = fees_q.filter(FeeRecord.company_id == selected_company_id)
        fees_rows = fees_q.group_by(db.extract('month', FeeRecord.payment_date)).all()
        fees_map = {int(r.m): float(r.total) for r in fees_rows}
        fees_monthly = [fees_map.get(m, 0.0) for m in range(1, 13)]
        ws = wb.active
        write_sheet(ws, 'Monthly Fees', ['Month'] + months, [['Collection (Rs.)'] + fees_monthly])
        
        rows = []
        course_fee_q = db.session.query(student_courses.c.course_id, db.func.sum(FeeRecord.amount_paid).label('total')).select_from(FeeRecord).join(Student).join(student_courses).join(Course, Course.id == student_courses.c.course_id).filter(FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date)
        if selected_company_id:
            course_fee_q = course_fee_q.filter(
                FeeRecord.company_id == selected_company_id,
                Course.company_id == selected_company_id
            )
        else:
            course_fee_q = course_fee_q.filter(Course.company_id == FeeRecord.company_id)
        course_fee_rows = course_fee_q.group_by(student_courses.c.course_id).all()
        course_fee_map = {r.course_id: float(r.total) for r in course_fee_rows}
        for course in Course.query.all():
            total = course_fee_map.get(course.id, 0.0)
            if total > 0:
                rows.append([course.name, course.code, total])
        if rows:
            ws2 = wb.create_sheet('Course-wise')
            write_sheet(ws2, 'Course-wise', ['Course', 'Code', 'Collected (Rs.)'], rows)

        daily_q = FeeRecord.query.filter(FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date)
        if selected_company_id:
            daily_q = daily_q.filter(FeeRecord.company_id == selected_company_id)
        daily = daily_q.order_by(FeeRecord.payment_date.desc()).all()
        if daily:
            ws3 = wb.create_sheet('Daily Collections')
            write_sheet(ws3, 'Daily Collections', ['Date', 'Student', 'Company', 'Taxable (Rs.)', 'GST (Rs.)', 'Total Paid (Rs.)', 'Method', 'Remarks'], [
                [r.payment_date.strftime('%d-%b-%Y'), r.student.name, r.company.name if r.company else 'Unassigned', r.taxable_amount, r.gst_amount, r.amount_paid, r.payment_method, r.remarks or ''] for r in daily
            ])

    elif tab == 'expense':
        months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        exp_q = db.session.query(db.extract('month', Expense.expense_date).label('m'), db.func.sum(Expense.amount).label('total')).filter(db.extract('year', Expense.expense_date) == filter_year)
        exp_q = filter_by_company_methods(exp_q, Expense.payment_method, selected_company_id)
        exp_rows = exp_q.group_by(db.extract('month', Expense.expense_date)).all()
        exp_map = {int(r.m): float(r.total) for r in exp_rows}
        monthly_exp = [exp_map.get(m, 0.0) for m in range(1, 13)]
        ws = wb.active
        write_sheet(ws, 'Monthly Expense', ['Month'] + months, [['Expense (Rs.)'] + monthly_exp])
        rows = []
        cat_exp_q = db.session.query(Expense.category_id, db.func.sum(Expense.amount).label('total')).filter(Expense.expense_date >= start_date, Expense.expense_date <= end_date)
        cat_exp_q = filter_by_company_methods(cat_exp_q, Expense.payment_method, selected_company_id)
        cat_exp_rows = cat_exp_q.group_by(Expense.category_id).all()
        cat_exp_map = {r.category_id: float(r.total) for r in cat_exp_rows}
        for cat in ExpenseCategory.query.all():
            total = cat_exp_map.get(cat.id, 0.0)
            rows.append([cat.name, total])
        ws2 = wb.create_sheet('Category-wise')
        write_sheet(ws2, 'Category-wise', ['Category', 'Total (Rs.)'], rows)
        recent_q = Expense.query.filter(Expense.expense_date >= start_date, Expense.expense_date <= end_date)
        recent_q = filter_by_company_methods(recent_q, Expense.payment_method, selected_company_id)
        recent = recent_q.order_by(Expense.expense_date.desc()).limit(50).all()
        if recent:
            ws3 = wb.create_sheet('Recent Expenses')
            write_sheet(ws3, 'Recent Expenses', ['Date', 'Category', 'Description', 'Amount'], [[r.expense_date.strftime('%d-%b-%Y'), r.category.name, r.description[:60], r.amount] for r in recent])

    elif tab == 'overall':
        months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        ws = wb.active
        tot_inc_q = db.session.query(db.func.sum(FeeRecord.amount_paid)).filter(FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date)
        gst_q = db.session.query(db.func.sum(FeeRecord.gst_amount)).filter(FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date)
        if selected_company_id:
            tot_inc_q = tot_inc_q.filter(FeeRecord.company_id == selected_company_id)
            gst_q = gst_q.filter(FeeRecord.company_id == selected_company_id)

        total_inc = tot_inc_q.scalar() or 0.0
        total_gst = gst_q.scalar() or 0.0
        
        total_exp_q = db.session.query(db.func.sum(Expense.amount)).filter(Expense.expense_date >= start_date, Expense.expense_date <= end_date)
        total_exp_q = filter_by_company_methods(total_exp_q, Expense.payment_method, selected_company_id)
        total_exp = total_exp_q.scalar() or 0.0
        
        total_fund_q = db.session.query(db.func.sum(OwnerFunding.amount)).filter(OwnerFunding.funding_date >= start_date, OwnerFunding.funding_date <= end_date)
        total_fund_q = filter_by_company_methods(total_fund_q, OwnerFunding.method, selected_company_id)
        total_fund = total_fund_q.scalar() or 0.0
        
        write_sheet(ws, 'P&L Summary', ['Metric', 'Value'], [
            ['Total Income', float(total_inc)],
            ['GST Collected', float(total_gst)],
            ['Capital Injection', float(total_fund)],
            ['Total Expense', float(total_exp)],
            ['Net Balance', float(total_inc) + float(total_fund) - float(total_exp)]
        ])
        
        rows = []
        inc_m_q = db.session.query(db.extract('month', FeeRecord.payment_date).label('m'), db.func.sum(FeeRecord.amount_paid).label('total')).filter(db.extract('year', FeeRecord.payment_date) == filter_year)
        if selected_company_id:
            inc_m_q = inc_m_q.filter(FeeRecord.company_id == selected_company_id)
        inc_rows_ov = inc_m_q.group_by(db.extract('month', FeeRecord.payment_date)).all()
        inc_map_ov = {int(r.m): float(r.total) for r in inc_rows_ov}
        
        exp_m_q = db.session.query(db.extract('month', Expense.expense_date).label('m'), db.func.sum(Expense.amount).label('total')).filter(db.extract('year', Expense.expense_date) == filter_year)
        exp_m_q = filter_by_company_methods(exp_m_q, Expense.payment_method, selected_company_id)
        exp_rows_ov = exp_m_q.group_by(db.extract('month', Expense.expense_date)).all()
        exp_map_ov = {int(r.m): float(r.total) for r in exp_rows_ov}
        
        fund_m_q = db.session.query(db.extract('month', OwnerFunding.funding_date).label('m'), db.func.sum(OwnerFunding.amount).label('total')).filter(db.extract('year', OwnerFunding.funding_date) == filter_year)
        fund_m_q = filter_by_company_methods(fund_m_q, OwnerFunding.method, selected_company_id)
        fund_rows_ov = fund_m_q.group_by(db.extract('month', OwnerFunding.funding_date)).all()
        fund_map_ov = {int(r.m): float(r.total) for r in fund_rows_ov}
        for m in range(1, 13):
            inc = inc_map_ov.get(m, 0.0)
            exp = exp_map_ov.get(m, 0.0)
            fund = fund_map_ov.get(m, 0.0)
            rows.append([months[m-1], inc, fund, exp, inc + fund - exp])
        ws2 = wb.create_sheet('Monthly P&L')
        write_sheet(ws2, 'Monthly P&L', ['Month', 'Income', 'Capital Injection', 'Expense', 'Net'], rows)

    elif tab == 'payment_methods':
        ws = wb.active
        ws.title = 'Summary'
        all_q = FeeRecord.query.filter(FeeRecord.payment_date >= start_date, FeeRecord.payment_date <= end_date)
        if selected_company_id:
            all_q = all_q.filter(FeeRecord.company_id == selected_company_id)
        all_records = all_q.order_by(FeeRecord.payment_date.desc()).all()
        total_period = sum(r.amount_paid for r in all_records)
        def filter_pm(records, account_name):
            matched = [r for r in records if classify_method(r.payment_method) == account_name]
            return sum(r.amount_paid for r in matched), matched
        summary_rows = []
        for label in PAYMENT_METHODS + ['Others']:
            total, records = filter_pm(all_records, label)
            if records:
                summary_rows.append([label, total])
                ws2 = wb.create_sheet(label[:20])
                write_sheet(ws2, label, ['Date', 'Student', 'Company', 'Amount', 'Remarks'], [[r.payment_date.strftime('%d-%b-%Y'), r.student.name, r.company.name if r.company else 'Unassigned', r.amount_paid, (r.remarks or '')] for r in records])
        write_sheet(ws, 'Payment Summary', ['Method', 'Total (Rs.)'], summary_rows)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filenames = {'income': 'Income_Report', 'fees': 'Student_Fees_Report', 'expense': 'Expense_Report', 'overall': 'Overall_Report', 'payment_methods': 'Payment_Methods_Report'}
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'{filenames.get(tab, "Report")}_{filter_year}_{filter_month}.xlsx')
